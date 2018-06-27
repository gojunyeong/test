import xml.etree.ElementTree as ET
import re
from openpyxl import Workbook
import datetime

templateParse = re.compile(
    r"(?P<template_key>[^=]*)=(?P<template_value>.*)\s\|")
getLinked = re.compile(r"[^[]*\[\[(([^\]]+)\|([^\]]*)|([^\]]*))\]\]")
getDateFunction = re.compile(
    r"[\w\W\s]+\|(\d*)\|(\d*)\|(\d*)(\|\w*=*\w*)?\}\}[\W\w\s]*")
getDateYMD = re.compile(r"\W*((\d+|)년|)\W*((\d+)월|)\W*((\d+)일|)")
getHeight = re.compile(r"\s*\D*(\d+)\D*")
ereaseRef = re.compile(r"<ref.*\/ref>")


tree = ET.parse('./input/kowiki-20180401-pages-articles-multistream.xml')
root = tree.getroot()


def makeXLSX(refinelist):
    wb = Workbook()
    ws = wb.active
    ws.title = "soccer player refined"
    ws['A1'] = "Page ID"
    ws['B1'] = "Page Title"
    ws['C1'] = "Template Name"
    ws['D1'] = "Date of Birth"
    ws['E1'] = "Height"
    ws['F1'] = "Team"
    for index, refineObject in enumerate(refinelist, 2):
        ws(1, index, refineObject["id"]+"")
        ws(2, index, refineObject["title"]+"")
        ws(3, index, refineObject["template_name"]+"")
        ws(4, index, refineObject["birthDate"]+"")
        ws(5, index, refineObject["height"]+"")
        ws(6, index, refineObject["currentTeam"]+"")

    wb.save('./output/soccer')


def refineBirthDate(date):

    date = ereaseRef.sub("", date)

    if date and date.strip():
        middate = ""
        if getDateFunction.search(date):
            midDate = getDateFunction.match(date)
            midDate = getDateFunction.sub(
                r"\g<1>-\g<2>-\g<3>", midDate.group())
        elif getDateYMD.search(date):
            midDate = getDateYMD.match(date)
            midDate = getDateYMD.sub(r"\g<2>-\g<4>-\g<6>", midDate.group())
        else:
            return date
        try:
            lastdate = datetime.datetime.strptime(
                midDate, "%Y-%m-%d").strftime('%Y-%m-%d')
            return lastdate
        except ValueError:
            lastdate = date.split('-')[0] and date.split('-')[0] or '0000' + date.split('-')[
                1] and date.split('-')[1] or '00' + date.split('-')[2] and date.split('-')[2] or '00'
            return lastdate

    else:
        return None


def refineValue(height):
    height = ereaseRef.sub("", height)
    print(height)
    if height and height.strip():
        midheight = float(getHeight.sub(r"\g<1>", height))
        return midheight < 5 and midheight*100 or midheight
    else:
        return None


def refineLinked(linked):
    linked = ereaseRef.sub("", linked)
    if linked and linked.strip():
        if getLinked.search(linked):
            midLinked = getLinked.match(linked)
            return getLinked.sub(r"\g<2>\g<4>", midLinked.group())
        else:
            return None
    else:
        return None


def getInfobox(text):
    startStatus = False
    startIndex = 0
    lastIndex = 0
    braketCount = 0
    template = list()
    infoBox = ''
    output = {"birthDate": None, "height": None, "currentTeam": None}
    for index, char in enumerate(text):
        if '{' == char:
            braketCount += 1
            if not startStatus:
                startIndex = index
                startStatus = True
        elif '}' == char:
            braketCount -= 1
            if startStatus and braketCount == 0:
                lastIndex = index+1
                template.append(text[startIndex:lastIndex])
                startStatus = False
    for text in template:
        if '축구 선수 정보' in text:
            infoBox = text
    tempList = templateParse.finditer(infoBox)
    for temp in tempList:
        if '출생일' in temp.group("template_key"):
            output["birthDate"] = refineBirthDate(temp.group("template_value"))
        elif '키' in temp.group("template_key"):
            output["height"] = refineValue(temp.group("template_value"))
        elif '현 소속팀' in temp.group("template_key"):
            output["currentTeam"] = refineLinked(temp.group("template_value"))

    return output


count = 0
outputList = list()
for page in root.findall('{http://www.mediawiki.org/xml/export-0.10/}page'):
    output = {}

    allText = page.find('{http://www.mediawiki.org/xml/export-0.10/}revision').find(
        '{http://www.mediawiki.org/xml/export-0.10/}text')

    if allText.text is not None and'축구 선수 정보' in allText.text:
        temp = getInfobox(allText.text)

        output["title"] = page.find(
            '{http://www.mediawiki.org/xml/export-0.10/}title').text
        print(output["title"])
        print(temp)
        output["id"] = page.find(
            '{http://www.mediawiki.org/xml/export-0.10/}id').text
        output["birthDate"] = temp["birthDate"]and temp["birthDate"] or None
        output["height"] = temp["height"] and temp["height"] or None
        output["currentTeam"] = temp["currentTeam"] and temp["currentTeam"] or None
        output["template_name"] = '축구 선수 정보'
        count = count+1
        outputList.append(output)
makeXLSX(outputList)
print(count)
